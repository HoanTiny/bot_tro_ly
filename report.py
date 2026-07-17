"""
Xuất báo cáo chi tiêu: file Excel (.xlsx) bằng openpyxl và biểu đồ PNG
bằng matplotlib.

Tách thành module riêng, không đụng Telegram hay Claude: nhận dữ liệu vào,
trả file ra — nhờ vậy test được offline (tạo file trong RAM rồi đọc lại).
"""

from io import BytesIO

import matplotlib

# Backend "Agg" = vẽ thẳng ra ảnh, không cần màn hình — bắt buộc khi chạy
# trên server/bot (backend mặc định sẽ đòi mở cửa sổ đồ họa và treo).
matplotlib.use("Agg")

import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font

from utils import format_money

HEADER_FONT = Font(bold=True)

# Màu biểu đồ: MỘT màu duy nhất cho các cột (đây là 1 chuỗi số liệu — tô mỗi
# nhóm một màu chỉ gây nhiễu), chữ màu xám đậm/nhạt cho phần phụ trợ.
# Biểu đồ so sánh 2 tuần: tuần này màu xanh (nổi), tuần trước xám (nền) —
# màu gắn với "tuần nào", không gắn với nhóm chi.
BAR_COLOR = "#2563eb"
LAST_WEEK_COLOR = "#c3cbd6"
TEXT_COLOR = "#111827"
MUTED_COLOR = "#6b7280"


def build_month_chart(summary: list[tuple[str, int]], year_month: str) -> BytesIO:
    """Vẽ biểu đồ cột ngang "chi theo nhóm" của 1 tháng, trả về PNG trong RAM.

    Chọn cột NGANG thay vì cột đứng/bánh tròn: tên nhóm tiếng Việt dài đọc
    thẳng hàng không phải xoay chữ, và so độ dài cột dễ hơn so góc miếng bánh.
    Số tiền ghi thẳng ở đầu mỗi cột nên bỏ luôn trục hoành cho sạch.

    summary: [(category, tổng), ...] nhóm lớn nhất trước (từ db.get_month_summary)
    """
    # barh vẽ từ dưới lên — đảo danh sách để nhóm chi nhiều nhất nằm TRÊN CÙNG
    categories = [category for category, _ in reversed(summary)]
    amounts = [amount for _, amount in reversed(summary)]

    # Cao ~0.5 inch mỗi cột: 2 nhóm hay 7 nhóm đều thoáng như nhau
    fig, ax = plt.subplots(figsize=(6.4, 0.5 * len(summary) + 1.1), dpi=150)
    bars = ax.barh(categories, amounts, color=BAR_COLOR, height=0.62)

    # Nhãn số tiền ở đầu cột — cách đầu cột 1.5% chiều dài cột lớn nhất
    gap = max(amounts) * 0.015
    for bar, amount in zip(bars, amounts):
        ax.text(
            bar.get_width() + gap, bar.get_y() + bar.get_height() / 2,
            format_money(amount), va="center", fontsize=10, color=TEXT_COLOR,
        )

    ax.set_title(
        f"Chi theo nhóm — tháng {year_month[5:7]}/{year_month[:4]}",
        fontsize=12, color=TEXT_COLOR, loc="left", pad=12,
    )
    # Dọn khung: bỏ 4 đường viền + trục hoành (nhãn số đã nói hết rồi),
    # chừa 18% bên phải cho nhãn số của cột dài nhất khỏi tràn ra ngoài ảnh
    for spine in ax.spines.values():
        spine.set_visible(False)
    ax.xaxis.set_visible(False)
    ax.tick_params(axis="y", length=0, labelsize=10, labelcolor=TEXT_COLOR)
    ax.set_xlim(0, max(amounts) * 1.18)
    fig.tight_layout()

    buffer = BytesIO()
    fig.savefig(buffer, format="png", facecolor="white")
    plt.close(fig)  # giải phóng bộ nhớ — matplotlib giữ figure mãi nếu không đóng
    buffer.seek(0)
    return buffer


def build_week_chart(
    this_week: list[tuple[str, int]],
    last_week: list[tuple[str, int]],
    this_label: str,
    last_label: str,
) -> BytesIO:
    """Vẽ biểu đồ so sánh chi theo nhóm giữa 2 tuần, trả về PNG trong RAM.

    Mỗi nhóm chi có 1 cặp cột ngang: tuần này (xanh) và tuần trước (xám nhạt) —
    nhìn phát biết ngay nhóm nào tăng/giảm. Nhóm chỉ có ở 1 trong 2 tuần
    vẫn hiện (cột kia bằng 0).

    this_week / last_week: [(category, tổng), ...] từ db.get_summary_between.
    this_label / last_label: chữ hiện trong chú giải, vd "10/07–16/07".
    """
    this_map, last_map = dict(this_week), dict(last_week)
    # Thứ tự nhóm: theo tuần này giảm dần, rồi các nhóm chỉ có ở tuần trước
    categories = [c for c, _ in this_week] + [c for c, _ in last_week if c not in this_map]

    fig, ax = plt.subplots(figsize=(6.4, 0.8 * len(categories) + 1.4), dpi=150)
    # barh vẽ từ dưới lên — đảo để nhóm chi nhiều nhất nằm trên cùng.
    # Mỗi nhóm chiếm 1 đơn vị trục dọc: cột tuần này nằm trên (+0.19),
    # cột tuần trước nằm dưới (-0.19), mỗi cột dày 0.34.
    positions = list(range(len(categories)))[::-1]
    this_vals = [this_map.get(c, 0) for c in categories]
    last_vals = [last_map.get(c, 0) for c in categories]

    bars_this = ax.barh(
        [p + 0.19 for p in positions], this_vals, height=0.34,
        color=BAR_COLOR, label=f"Tuần này ({this_label})",
    )
    bars_last = ax.barh(
        [p - 0.19 for p in positions], last_vals, height=0.34,
        color=LAST_WEEK_COLOR, label=f"Tuần trước ({last_label})",
    )

    # Nhãn số tiền ở đầu mỗi cột (cột 0đ thì bỏ nhãn cho đỡ rối)
    biggest = max(this_vals + last_vals)
    gap = biggest * 0.015
    for bars, color in ((bars_this, TEXT_COLOR), (bars_last, MUTED_COLOR)):
        for bar in bars:
            if bar.get_width() > 0:
                ax.text(
                    bar.get_width() + gap, bar.get_y() + bar.get_height() / 2,
                    format_money(int(bar.get_width())), va="center",
                    fontsize=9, color=color,
                )

    ax.set_yticks(positions)
    ax.set_yticklabels(categories)
    ax.set_title("Chi theo nhóm — tuần này so với tuần trước",
                  fontsize=12, color=TEXT_COLOR, loc="left", pad=12)
    ax.legend(loc="lower right", frameon=False, fontsize=9)
    for spine in ax.spines.values():
        spine.set_visible(False)
    ax.xaxis.set_visible(False)
    ax.tick_params(axis="y", length=0, labelsize=10, labelcolor=TEXT_COLOR)
    ax.set_xlim(0, biggest * 1.18)
    fig.tight_layout()

    buffer = BytesIO()
    fig.savefig(buffer, format="png", facecolor="white")
    plt.close(fig)
    buffer.seek(0)
    return buffer


def build_month_report(
    rows: list[tuple[str, int, str, str, str]],
    summary: list[tuple[str, int]],
    year_month: str,
    total_income: int = 0,
) -> BytesIO:
    """Tạo file Excel 2 sheet: "Chi tiết" (từng khoản) và "Tổng hợp" (theo nhóm).

    rows: [(item, amount, category, "dd/mm", kind), ...] — mới nhất trước (từ db)
    total_income: tổng tiền thu trong tháng (hiện ở sheet Tổng hợp)
    Trả về BytesIO — file nằm trong RAM, gửi thẳng qua Telegram không cần
    ghi ra đĩa (đỡ phải dọn file tạm).
    """
    wb = Workbook()

    # ── Sheet 1: từng khoản thu/chi, cũ -> mới cho dễ đọc ────────────────
    ws = wb.active
    ws.title = "Chi tiết"
    ws.append(["Ngày", "Khoản", "Nhóm", "Loại", "Số tiền (đ)"])
    for cell in ws[1]:
        cell.font = HEADER_FONT

    for item, amount, category, day, kind in reversed(rows):
        ws.append([day, item, category, "Thu" if kind == "thu" else "Chi", amount])

    # Định dạng cột tiền có ngăn cách hàng nghìn (Excel tự hiển thị 15,000)
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=5).number_format = "#,##0"

    # Nới độ rộng cột cho dễ nhìn (đơn vị ~ số ký tự)
    for col, width in zip("ABCDE", (10, 30, 14, 8, 14)):
        ws.column_dimensions[col].width = width

    # ── Sheet 2: tổng theo nhóm + thu, chi, cân đối ──────────────────────
    ws2 = wb.create_sheet("Tổng hợp")
    ws2.append([f"Thu chi tháng {year_month[5:7]}/{year_month[:4]}", ""])
    ws2["A1"].font = HEADER_FONT
    ws2.append(["Nhóm", "Tổng (đ)"])
    for cell in ws2[2]:
        cell.font = HEADER_FONT

    for category, total in summary:
        ws2.append([category, total])

    total_expense = sum(total for _, total in summary)
    for label, value in (
        ("TỔNG CHI", total_expense),
        ("TỔNG THU", total_income),
        ("CÂN ĐỐI", total_income - total_expense),
    ):
        ws2.append([label, value])
        for cell in ws2[ws2.max_row]:
            cell.font = HEADER_FONT

    for row_idx in range(3, ws2.max_row + 1):
        ws2.cell(row=row_idx, column=2).number_format = "#,##0"
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 14

    # Lưu vào RAM thay vì file trên đĩa
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)  # tua con trỏ về đầu để bên nhận đọc từ byte đầu tiên
    return buffer
