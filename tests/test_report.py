"""Test xuất Excel: tạo file trong RAM rồi đọc ngược lại kiểm tra nội dung."""

from openpyxl import load_workbook

from report import build_month_report

ROWS = [  # mới nhất trước, như db trả về: (item, amount, category, day, kind)
    ("nhận lương", 15000000, "thu nhập", "17/07", "thu"),
    ("đổ xăng", 50000, "đi lại", "16/07", "chi"),
    ("ăn sáng", 15000, "ăn uống", "15/07", "chi"),
]
SUMMARY = [("ăn uống", 15000), ("đi lại", 50000)]


def test_bao_cao_co_du_2_sheet_va_dung_so_lieu():
    buffer = build_month_report(ROWS, SUMMARY, "2026-07", total_income=15000000)
    wb = load_workbook(buffer)

    assert wb.sheetnames == ["Chi tiết", "Tổng hợp"]

    # Sheet chi tiết: header + 3 dòng, đảo lại thành cũ -> mới, có cột Loại
    ws = wb["Chi tiết"]
    assert [c.value for c in ws[1]] == ["Ngày", "Khoản", "Nhóm", "Loại", "Số tiền (đ)"]
    assert [c.value for c in ws[2]] == ["15/07", "ăn sáng", "ăn uống", "Chi", 15000]
    assert [c.value for c in ws[4]] == ["17/07", "nhận lương", "thu nhập", "Thu", 15000000]

    # Sheet tổng hợp: 3 dòng cuối là TỔNG CHI / TỔNG THU / CÂN ĐỐI
    ws2 = wb["Tổng hợp"]
    last3 = [[c.value for c in ws2[r]] for r in range(ws2.max_row - 2, ws2.max_row + 1)]
    assert last3 == [
        ["TỔNG CHI", 65000],
        ["TỔNG THU", 15000000],
        ["CÂN ĐỐI", 15000000 - 65000],
    ]


def test_bao_cao_thang_khong_co_khoan_nao():
    # Danh sách rỗng vẫn phải tạo được file hợp lệ (handler đã chặn trước,
    # nhưng hàm không nên crash nếu bị gọi với dữ liệu rỗng)
    buffer = build_month_report([], [], "2026-01")
    wb = load_workbook(buffer)
    assert wb["Tổng hợp"]["A3"].value == "TỔNG CHI"
    assert wb["Tổng hợp"]["B3"].value == 0
    assert wb["Tổng hợp"]["A5"].value == "CÂN ĐỐI"
